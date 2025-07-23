import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import json
from datetime import datetime
import os
import jdatetime  # برای کار با تاریخ شمسی

# برای Excel
try:
    import openpyxl
    from openpyxl.styles import PatternFill
    from openpyxl.utils import get_column_letter

    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("OpenPyXL not installed. Excel export disabled.")

# برای PDF
try:
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False
    print("ReportLab not installed. PDF export disabled.")

# ---------- تنظیمات فونت فارسی برای PDF و UI (مهم) ----------
# نام فونت فارسی برای UI Tkinter و PDF
GLOBAL_FONT_NAME = "Tanha"  # مطمئن شوید فایل Tanha.ttf در کنار برنامه هست
GLOBAL_FONT_SIZE = 10  # اندازه فونت کلی
PDF_FONT_PATH = "Tanha.ttf"
PDF_FONT_NAME = "Tanha"


def register_persian_font_for_pdf():
    """ثبت فونت فارسی برای استفاده در ReportLab."""
    if not PDF_AVAILABLE:
        return

    if not os.path.exists(PDF_FONT_PATH):
        messagebox.showwarning("هشدار فونت PDF",
                               f"فایل فونت PDF ({PDF_FONT_PATH}) یافت نشد. "
                               "لطفاً آن را دانلود کرده و در کنار برنامه قرار دهید."
                               "\nگزارش PDF ممکن است متون فارسی را به درستی نمایش ندهد.")
        return

    try:
        # فقط در صورتی که فونت قبلاً ثبت نشده باشد، آن را ثبت کنید
        if PDF_FONT_NAME not in pdfmetrics.getRegisteredFontNames():
            pdfmetrics.registerFont(TTFont(PDF_FONT_NAME, PDF_FONT_PATH))
    except Exception as e:
        messagebox.showerror("خطای فونت PDF", f"خطا در بارگذاری فونت PDF: {e}")


# ---------- توابع تبدیل تاریخ شمسی/میلادی ----------
def parse_shamsi_date(sh_date_str):
    """
    تبدیل رشته تاریخ شمسی (YYYY/MM/DD) به شیء jdatetime.date.
    در صورت عدم موفقیت None برمی‌گرداند.
    """
    try:
        if not sh_date_str:
            return None
        return jdatetime.datetime.strptime(sh_date_str, '%Y/%m/%d').date()
    except ValueError:
        return None


def shamsi_to_gregorian_datetime(sh_date_str):
    """
    تبدیل رشته تاریخ شمسی به شیء datetime میلادی.
    در صورت عدم موفقیت None برمی‌گرداند.
    """
    jdate = parse_shamsi_date(sh_date_str)
    if jdate:
        return jdate.togregorian()
    return None


def gregorian_datetime_to_shamsi_str(dt_obj):
    """
    تبدیل شیء datetime میلادی به رشته تاریخ شمسی (YYYY/MM/DD).
    """
    if not dt_obj:
        return ""
    jdate = jdatetime.date.fromgregorian(date=dt_obj)
    return jdate.strftime('%Y/%m/%d')


# ---------- ذخیره و بارگذاری داده‌ها و تنظیمات ----------
DATA_FILE = "projects_data.json"
CONFIG_FILE = "config.json"


def save_data(data):
    """ذخیره داده‌ها در فایل JSON"""
    try:
        with open(DATA_FILE, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    except Exception as e:
        messagebox.showerror("خطا", f"خطا در ذخیره داده‌ها: {str(e)}")


def load_data():
    """بارگذاری داده‌ها از فایل JSON"""
    try:
        if not os.path.exists(DATA_FILE):
            return []
        with open(DATA_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception as e:
        messagebox.showerror("خطا", f"خطا در بارگذاری داده‌ها: {str(e)}")
        return []


def save_config(config):
    """ذخیره تنظیمات در فایل JSON"""
    try:
        with open(CONFIG_FILE, "w", encoding="utf-8") as f:
            json.dump(config, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"Error saving config: {e}")


def load_config():
    """بارگذاری تنظیمات از فایل JSON"""
    try:
        if not os.path.exists(CONFIG_FILE):
            return {"theme": "light"}  # تم پیش‌فرض
        with open(CONFIG_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception as e:
        print(f"Error loading config: {e}")
        return {"theme": "light"}


# ---------- تعیین وضعیت ----------
def determine_status(next_call_date_str, finished):
    """تعیین وضعیت پروژه بر اساس تاریخ تماس بعدی"""
    if finished:
        return ""

    if not next_call_date_str:
        return "انتظار"

    dt_today = datetime.now().date()
    dt_next_call = shamsi_to_gregorian_datetime(next_call_date_str)

    if dt_next_call is None:
        return "انتظار"
    elif dt_next_call > dt_today:
        return "انتظار"
    else:
        return "در انتظار تماس مجدد"


# ---------- کلاس اصلی برنامه ----------
class ProjectManager:
    def __init__(self, root):
        self.root = root
        self.root.title("مدیریت پروژه‌ها")
        self.root.geometry("1400x800")
        self.root.minsize(1000, 600)

        # تنظیم فونت سراسری برای Tkinter widgets (غیر-ttk)
        root.option_add("*Font", (GLOBAL_FONT_NAME, GLOBAL_FONT_SIZE))

        self.style = ttk.Style(self.root)
        self.config = load_config()
        self.current_theme = self.config.get("theme", "light")

        # داده‌ها
        self.data = load_data()

        # متغیرها
        self.entries = {}
        self.finished_var = tk.BooleanVar()
        self.finished_status_var = tk.StringVar()

        # فیلترها
        self.filter_status_var = tk.StringVar()
        self.filter_name_var = tk.StringVar()
        self.filter_keyword_var = tk.StringVar()
        self.filter_date_from_var = tk.StringVar()
        self.filter_date_to_var = tk.StringVar()

        # مرتب‌سازی
        self.sort_by_var = tk.StringVar()
        self.sort_order_var = tk.StringVar()

        self.create_widgets()
        self.apply_theme(self.current_theme)
        self.refresh_table()
        self.update_status_bar("برنامه آماده است.")

    def create_widgets(self):
        """ایجاد عناصر واسط کاربری"""
        toolbar_frame = ttk.Frame(self.root, padding="5 5 5 5")
        toolbar_frame.pack(side="top", fill="x", padx=10, pady=(5, 0))

        self.theme_toggle_button = ttk.Button(toolbar_frame, text="حالت تاریک", command=self.toggle_theme)
        self.theme_toggle_button.pack(side="left")

        main_frame = ttk.Frame(self.root, padding="10 10 10 10")
        main_frame.pack(fill="both", expand=True, padx=10, pady=5)

        self.create_form(main_frame)
        self.create_buttons(main_frame)
        self.create_filter_sort(main_frame)
        self.create_table(main_frame)
        self.create_export_buttons(main_frame)

        self.status_bar = ttk.Label(self.root, text="", relief=tk.SUNKEN, anchor="w", padding="5 0 0 0",style="Statusbar.TLabel")
        self.status_bar.pack(side="bottom", fill="x")

    def create_form(self, parent_frame):
        """ایجاد فرم ورود داده"""
        frame_form = ttk.LabelFrame(parent_frame, text="ورود اطلاعات پروژه", padding="10")
        frame_form.pack(padx=5, pady=5, fill="x")

        # ردیف ۰
        row0_frame = ttk.Frame(frame_form)
        row0_frame.pack(fill="x", pady=2)
        ttk.Label(row0_frame, text="نام مهندس:").pack(side="right", padx=5)
        self.entries["name"] = ttk.Entry(row0_frame, width=25, justify="right")
        self.entries["name"].pack(side="right", padx=5, expand=True, fill="x")
        ttk.Label(row0_frame, text="آدرس:").pack(side="right", padx=(20, 5))
        self.entries["address"] = ttk.Entry(row0_frame, width=40, justify="right")
        self.entries["address"].pack(side="right", padx=5, expand=True, fill="x")

        # ردیف ۱
        row1_frame = ttk.Frame(frame_form)
        row1_frame.pack(fill="x", pady=2)
        ttk.Label(row1_frame, text="متراژ:").pack(side="right", padx=5)
        self.entries["area"] = ttk.Entry(row1_frame, width=25, justify="right")
        self.entries["area"].pack(side="right", padx=5, expand=True, fill="x")
        ttk.Label(row1_frame, text="تعداد اتاق:").pack(side="right", padx=(20, 5))
        self.entries["rooms"] = ttk.Entry(row1_frame, width=25, justify="right")
        self.entries["rooms"].pack(side="right", padx=5, expand=True, fill="x")

        # ردیف ۲
        row2_frame = ttk.Frame(frame_form)
        row2_frame.pack(fill="x", pady=2)
        ttk.Label(row2_frame, text="تاریخ ویزیت (۱۴۰۲/۰۱/۰۱):").pack(side="right", padx=5)
        self.entries["visit_date"] = ttk.Entry(row2_frame, width=25, justify="right")
        self.entries["visit_date"].pack(side="right", padx=5, expand=True, fill="x")
        ttk.Label(row2_frame, text="تاریخ تماس بعدی (اختیاری):").pack(side="right", padx=(20, 5))
        self.entries["next_call_date"] = ttk.Entry(row2_frame, width=25, justify="right")
        self.entries["next_call_date"].pack(side="right", padx=5, expand=True, fill="x")

        # ردیف ۳ - توضیحات
        row3_frame = ttk.Frame(frame_form)
        row3_frame.pack(fill="x", pady=2)
        ttk.Label(row3_frame, text="توضیحات:").pack(side="right", padx=5, anchor="ne")
        # فونت برای tk.Text باید به صورت جداگانه تنظیم شود
        self.entries["description"] = tk.Text(row3_frame, height=3, width=50, wrap="word",
                                              font=(GLOBAL_FONT_NAME, GLOBAL_FONT_SIZE))
        self.entries["description"].pack(side="right", padx=5, expand=True, fill="x")

        # ردیف ۴ - وضعیت تمام شده
        finished_frame = ttk.Frame(frame_form)
        finished_frame.pack(fill="x", pady=5)

        self.entries["end_date"] = ttk.Entry(finished_frame, width=15, justify="right")
        self.entries["end_date"].pack(side="right", padx=(5, 0))

        ttk.Label(finished_frame, text="تاریخ پایان:").pack(side="right", padx=(20, 5))

        self.finished_status_dropdown = ttk.Combobox(finished_frame, textvariable=self.finished_status_var,
                                                     state="disabled", width=15, justify="right")
        self.finished_status_dropdown['values'] = ("از دست رفته", "خرید")
        self.finished_status_dropdown.pack(side="right", padx=(5, 0))

        ttk.Label(finished_frame, text="وضعیت:").pack(side="right", padx=(20, 5))

        self.finished_check = ttk.Checkbutton(finished_frame, text="تمام شده", variable=self.finished_var)
        self.finished_check.pack(side="right")

        self.finished_var.trace_add("write", self.on_finished_change)

    def create_buttons(self, parent_frame):
        """ایجاد دکمه‌های عملیات اصلی"""
        frame_buttons = ttk.Frame(parent_frame, padding="5")
        frame_buttons.pack(padx=5, pady=5, fill="x")

        ttk.Button(frame_buttons, text="افزودن/ویرایش", command=self.add_or_update_entry,
                   style="Accent.TButton").pack(side="right", padx=5)
        ttk.Button(frame_buttons, text="حذف انتخاب شده", command=self.delete_selected,
                   style="Danger.TButton").pack(side="right", padx=5)
        ttk.Button(frame_buttons, text="پاک کردن فرم", command=self.clear_fields).pack(side="right", padx=5)
        ttk.Button(frame_buttons, text="بارگذاری در فرم", command=self.load_to_form).pack(side="right", padx=5)

    def create_filter_sort(self, parent_frame):
        """ایجاد بخش فیلتر و مرتب‌سازی"""
        frame_filter = ttk.LabelFrame(parent_frame, text="فیلتر و مرتب‌سازی", padding="10")
        frame_filter.pack(padx=5, pady=5, fill="x")

        filter_row1 = ttk.Frame(frame_filter)
        filter_row1.pack(fill="x", pady=2)

        ttk.Button(filter_row1, text="اعمال", command=self.apply_filter_sort, style="Primary.TButton").pack(side="left",
                                                                                                            padx=5)
        ttk.Button(filter_row1, text="پاک کردن فیلتر", command=self.clear_filters).pack(side="left", padx=5)

        self.sort_order = ttk.Combobox(filter_row1, textvariable=self.sort_order_var, width=10, justify="right",
                                       state="readonly")
        self.sort_order['values'] = ("صعودی", "نزولی")
        self.sort_order.set("صعودی")
        self.sort_order.pack(side="right", padx=5)

        self.sort_by = ttk.Combobox(filter_row1, textvariable=self.sort_by_var, width=18, justify="right",
                                    state="readonly")
        self.sort_by['values'] = ("تاریخ تماس بعدی", "تاریخ ویزیت", "تاریخ پایان", "نام مهندس", "وضعیت")
        self.sort_by.set("تاریخ تماس بعدی")
        self.sort_by.pack(side="right", padx=5)
        ttk.Label(filter_row1, text="مرتب‌سازی:").pack(side="right", padx=(20, 5))

        filter_row2 = ttk.Frame(frame_filter)
        filter_row2.pack(fill="x", pady=2)

        self.filter_keyword = ttk.Entry(filter_row2, textvariable=self.filter_keyword_var, width=20, justify="right")
        self.filter_keyword.pack(side="right", padx=5)
        ttk.Label(filter_row2, text="کلیدواژه (توضیحات):").pack(side="right", padx=(20, 5))

        self.filter_name = ttk.Entry(filter_row2, textvariable=self.filter_name_var, width=20, justify="right")
        self.filter_name.pack(side="right", padx=5)
        ttk.Label(filter_row2, text="مهندس:").pack(side="right", padx=(20, 5))

        self.filter_status = ttk.Combobox(filter_row2, textvariable=self.filter_status_var, width=15, justify="right",
                                          state="readonly")
        self.filter_status['values'] = ("همه", "از دست رفته", "خرید", "انتظار", "در انتظار تماس مجدد")
        self.filter_status.set("همه")
        self.filter_status.pack(side="right", padx=5)
        ttk.Label(filter_row2, text="وضعیت:").pack(side="right", padx=(5, 0))

        filter_row3 = ttk.Frame(frame_filter)
        filter_row3.pack(fill="x", pady=2)

        self.filter_date_to = ttk.Entry(filter_row3, textvariable=self.filter_date_to_var, width=15, justify="right")
        self.filter_date_to.pack(side="right", padx=5)
        ttk.Label(filter_row3, text="تا (تاریخ تماس بعدی):").pack(side="right", padx=(5, 5))

        self.filter_date_from = ttk.Entry(filter_row3, textvariable=self.filter_date_from_var, width=15,
                                          justify="right")
        self.filter_date_from.pack(side="right", padx=5)
        ttk.Label(filter_row3, text="تاریخ از:").pack(side="right", padx=(20, 5))

    def create_table(self, parent_frame):
        """ایجاد جدول نمایش داده‌ها"""
        table_frame = ttk.Frame(parent_frame, padding="5")
        table_frame.pack(padx=5, pady=5, fill="both", expand=True)

        cols = ("نام مهندس", "آدرس", "متراژ", "تعداد اتاق", "تاریخ ویزیت",
                "تاریخ تماس بعدی", "وضعیت", "توضیحات", "تاریخ پایان")

        self.tree = ttk.Treeview(table_frame, columns=cols, show="headings", height=15)

        column_widths = {
            "نام مهندس": 120, "آدرس": 200, "متراژ": 80, "تعداد اتاق": 100,
            "تاریخ ویزیت": 120, "تاریخ تماس بعدی": 120, "وضعیت": 150,
            "توضیحات": 200, "تاریخ پایان": 120
        }
        for col in cols:
            self.tree.heading(col, text=col)
            # فونت برای سربرگ Treeview
            self.tree.column(col, width=column_widths.get(col, 100), anchor="center")

        v_scrollbar = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        h_scrollbar = ttk.Scrollbar(table_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)

        self.tree.pack(side="left", fill="both", expand=True)
        v_scrollbar.pack(side="right", fill="y")
        h_scrollbar.pack(side="bottom", fill="x")

    def create_export_buttons(self, parent_frame):
        """ایجاد دکمه‌های خروجی"""
        export_frame = ttk.Frame(parent_frame, padding="5")
        export_frame.pack(padx=5, pady=5, fill="x")

        if EXCEL_AVAILABLE:
            ttk.Button(export_frame, text="خروجی Excel", command=self.export_to_excel,
                       style="Success.TButton").pack(side="right", padx=5)
        else:
            ttk.Button(export_frame, text="Excel غیرفعال (نیاز به openpyxl)",
                       state="disabled").pack(side="right", padx=5)

        if PDF_AVAILABLE:
            ttk.Button(export_frame, text="خروجی PDF", command=self.export_to_pdf,
                       style="Info.TButton").pack(side="right", padx=5)
            register_persian_font_for_pdf()
        else:
            ttk.Button(export_frame, text="PDF غیرفعال (نیاز به reportlab)",
                       state="disabled").pack(side="right", padx=5)

    def on_finished_change(self, *args):
        """رویداد تغییر وضعیت تمام شده"""
        if self.finished_var.get():
            self.finished_status_dropdown.config(state="readonly")
            if not self.entries["end_date"].get():
                self.entries["end_date"].insert(0, gregorian_datetime_to_shamsi_str(datetime.now().date()))
        else:
            self.finished_status_var.set("")
            self.finished_status_dropdown.config(state="disabled")
            self.entries["end_date"].delete(0, tk.END)

    def clear_fields(self):
        """پاک کردن فیلدهای فرم"""
        for key, widget in self.entries.items():
            if isinstance(widget, ttk.Entry) or isinstance(widget, tk.Entry):
                widget.delete(0, tk.END)
            elif isinstance(widget, tk.Text):
                widget.delete("1.0", tk.END)
        self.finished_var.set(False)
        self.finished_status_var.set("")
        self.update_status_bar("فرم پاک شد.")

    def clear_filters(self):
        """پاک کردن فیلترها"""
        self.filter_status_var.set("همه")
        self.filter_name_var.set("")
        self.filter_keyword_var.set("")
        self.filter_date_from_var.set("")
        self.filter_date_to_var.set("")
        self.refresh_table()
        self.update_status_bar("فیلترها پاک شدند.")

    def load_to_form(self):
        """بارگذاری رکورد انتخاب شده در فرم"""
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning("اخطار", "لطفاً یک رکورد انتخاب کنید.")
            return

        item = self.tree.item(selected[0])
        values = item["values"]
        if not values:
            return

        selected_name = values[0]
        selected_address = values[1]

        found_rec = None
        for rec in self.data:
            if rec.get("name") == selected_name and rec.get("address") == selected_address:
                found_rec = rec
                break

        if found_rec:
            self.clear_fields()

            self.entries["name"].insert(0, found_rec.get("name", ""))
            self.entries["address"].insert(0, found_rec.get("address", ""))
            self.entries["area"].insert(0, found_rec.get("area", ""))
            self.entries["rooms"].insert(0, found_rec.get("rooms", ""))
            self.entries["visit_date"].insert(0, found_rec.get("visit_date", ""))
            self.entries["next_call_date"].insert(0, found_rec.get("next_call_date", ""))

            description_text = found_rec.get("description", "")
            self.entries["description"].insert("1.0", description_text)

            end_date_text = found_rec.get("end_date", "")
            self.entries["end_date"].insert(0, end_date_text)

            status = found_rec.get("status", "")
            if status in ("از دست رفته", "خرید"):
                self.finished_var.set(True)
                self.finished_status_var.set(status)
            else:
                self.finished_var.set(False)
                self.finished_status_var.set("")
            self.update_status_bar(f"رکورد '{selected_name}' در فرم بارگذاری شد.")
        else:
            messagebox.showerror("خطا", "رکورد یافت نشد. ممکن است داده‌ها تغییر کرده باشند.")
            self.update_status_bar("خطا: رکورد یافت نشد.")

    def refresh_table(self, filtered_data=None):
        """بروزرسانی جدول"""
        for row in self.tree.get_children():
            self.tree.delete(row)

        display_data = filtered_data if filtered_data is not None else self.data

        self.style_treeview_tags()

        for rec in display_data:
            current_status = rec.get("status", "")
            is_finished_in_data = (current_status in ("از دست رفته", "خرید"))
            if not is_finished_in_data:
                rec["status"] = determine_status(rec.get("next_call_date"), is_finished_in_data)

            vals = (
                rec.get("name", ""),
                rec.get("address", ""),
                rec.get("area", ""),
                rec.get("rooms", ""),
                rec.get("visit_date", ""),
                rec.get("next_call_date", ""),
                rec.get("status", ""),
                rec.get("description", "")[:50] + "..." if len(rec.get("description", "")) > 50 else rec.get(
                    "description", ""),
                rec.get("end_date", "")
            )

            status = rec.get("status", "")
            tag = ""
            if status == "از دست رفته":
                tag = "tag_red"
            elif status == "خرید":
                tag = "tag_green"
            elif status == "انتظار":
                tag = "tag_yellow"
            elif status == "در انتظار تماس مجدد":
                tag = "tag_blue"

            self.tree.insert("", "end", values=vals, tags=(tag,))

    def add_or_update_entry(self):
        """افزودن یا ویرایش رکورد"""
        name = self.entries["name"].get().strip()
        address = self.entries["address"].get().strip()
        area = self.entries["area"].get().strip()
        rooms = self.entries["rooms"].get().strip()
        visit_date = self.entries["visit_date"].get().strip()
        next_call_date = self.entries["next_call_date"].get().strip()
        description = self.entries["description"].get("1.0", tk.END).strip()
        finished = self.finished_var.get()
        status = self.finished_status_var.get() if finished else ""
        end_date = self.entries["end_date"].get().strip() if finished else ""

        if not name or not visit_date:
            messagebox.showerror("خطا", "لطفاً حداقل نام مهندس و تاریخ ویزیت را وارد کنید.")
            self.update_status_bar("خطا: نام مهندس یا تاریخ ویزیت خالی است.")
            return

        if parse_shamsi_date(visit_date) is None:
            messagebox.showerror("خطا", "فرمت تاریخ ویزیت صحیح نیست (مثال: ۱۴۰۲/۰۱/۰۱).")
            self.update_status_bar("خطا: فرمت تاریخ ویزیت اشتباه است.")
            return

        if next_call_date and parse_shamsi_date(next_call_date) is None:
            messagebox.showerror("خطا", "فرمت تاریخ تماس بعدی صحیح نیست (مثال: ۱۴۰۲/۰۱/۰۱).")
            self.update_status_bar("خطا: فرمت تاریخ تماس بعدی اشتباه است.")
            return

        if finished and end_date and parse_shamsi_date(end_date) is None:
            messagebox.showerror("خطا", "فرمت تاریخ پایان صحیح نیست (مثال: ۱۴۰۲/۰۱/۰۱).")
            self.update_status_bar("خطا: فرمت تاریخ پایان اشتباه است.")
            return

        actual_status = status if finished else determine_status(next_call_date, finished)

        found = False
        for rec in self.data:
            if rec.get("name") == name and rec.get("address") == address:
                rec.update({
                    "area": area,
                    "rooms": rooms,
                    "visit_date": visit_date,
                    "next_call_date": next_call_date,
                    "status": actual_status,
                    "description": description,
                    "end_date": end_date
                })
                found = True
                break

        if not found:
            new_rec = {
                "name": name,
                "address": address,
                "area": area,
                "rooms": rooms,
                "visit_date": visit_date,
                "next_call_date": next_call_date,
                "status": actual_status,
                "description": description,
                "end_date": end_date
            }
            self.data.append(new_rec)

        save_data(self.data)
        self.refresh_table()
        self.clear_fields()
        self.update_status_bar("رکورد با موفقیت ذخیره شد.")

    def delete_selected(self):
        """حذف رکورد انتخاب شده"""
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning("اخطار", "لطفاً یک رکورد برای حذف انتخاب کنید.")
            self.update_status_bar("اخطار: رکوردی برای حذف انتخاب نشده.")
            return

        if messagebox.askyesno("تایید حذف", "آیا مطمئن هستید که می‌خواهید این رکورد را حذف کنید؟"):
            item = self.tree.item(selected[0])
            values = item["values"]
            if not values:
                return

            name = values[0]
            address = values[1]
            self.data = [rec for rec in self.data if not (rec.get("name") == name and rec.get("address") == address)]
            save_data(self.data)
            self.refresh_table()
            self.update_status_bar("رکورد با موفقیت حذف شد.")

    def apply_filter_sort(self):
        """اعمال فیلتر و مرتب‌سازی"""
        filtered = []
        status_filter = self.filter_status_var.get()
        name_filter = self.filter_name_var.get().strip().lower()
        keyword_filter = self.filter_keyword_var.get().strip().lower()
        date_from_str = self.filter_date_from_var.get().strip()
        date_to_str = self.filter_date_to_var.get().strip()

        dt_from = shamsi_to_gregorian_datetime(date_from_str) if date_from_str else None
        dt_to = shamsi_to_gregorian_datetime(date_to_str) if date_to_str else None

        for rec in self.data:
            if status_filter != "همه" and rec.get("status") != status_filter:
                continue

            if name_filter and name_filter not in rec.get("name", "").lower():
                continue

            if keyword_filter and keyword_filter not in rec.get("description", "").lower():
                continue

            next_call_date_rec = rec.get("next_call_date", "")
            if next_call_date_rec:
                dt_next_call_rec = shamsi_to_gregorian_datetime(next_call_date_rec)
                if dt_next_call_rec:
                    if dt_from and dt_next_call_rec < dt_from:
                        continue
                    if dt_to and dt_next_call_rec > dt_to:
                        continue
                elif dt_from or dt_to:
                    continue
            elif dt_from or dt_to:
                continue

            filtered.append(rec)

        sort_by = self.sort_by_var.get()
        sort_order = self.sort_order_var.get()
        reverse = (sort_order == "نزولی")

        def get_sort_key(rec):
            if sort_by == "تاریخ تماس بعدی":
                dt = shamsi_to_gregorian_datetime(rec.get("next_call_date", ""))
                return dt if dt is not None else datetime.min
            elif sort_by == "تاریخ ویزیت":
                dt = shamsi_to_gregorian_datetime(rec.get("visit_date", ""))
                return dt if dt is not None else datetime.min
            elif sort_by == "تاریخ پایان":
                dt = shamsi_to_gregorian_datetime(rec.get("end_date", ""))
                return dt if dt is not None else datetime.min
            elif sort_by == "نام مهندس":
                return rec.get("name", "")
            elif sort_by == "وضعیت":
                status_order = {"در انتظار تماس مجدد": 1, "انتظار": 2, "خرید": 3, "از دست رفته": 4, "": 5}
                return status_order.get(rec.get("status", ""), 99)
            return ""

        filtered.sort(key=get_sort_key, reverse=reverse)
        self.refresh_table(filtered)
        self.update_status_bar(f"{len(filtered)} رکورد فیلتر و مرتب‌سازی شد.")

    def update_status_bar(self, message, duration_ms=3000):
        """نمایش پیام در نوار وضعیت"""
        self.status_bar.config(text=message)
        self.root.after(duration_ms, lambda: self.status_bar.config(text=""))

    def apply_theme(self, theme_name):
        """اعمال تم (روشن یا تاریک) به تمام عناصر UI"""
        self.current_theme = "light"

        if theme_name == "dark":
            bg_color = "#2e2e2e"
            fg_color = "#ffffff"
            entry_bg = "#3a3a3a"
            entry_fg = "#ffffff"
            select_bg = "#007acc"
            select_fg = "#ffffff"

            btn_default_bg = "#555555"
            btn_hover_bg = "#666666"
            btn_accent_bg = "#4CAF50"
            btn_danger_bg = "#E53935"
            btn_primary_bg = "#2196F3"
            btn_info_bg = "#00BCD4"

            self.theme_toggle_button.config(text="حالت روشن")

        else:  # light
            bg_color = "#f0f0f0"
            fg_color = "#333333"
            entry_bg = "#ffffff"
            entry_fg = "#000000"
            select_bg = "#bde2ff"
            select_fg = "#000000"

            btn_default_bg = "#e1e1e1"
            btn_hover_bg = "#d1d1d1"
            btn_accent_bg = "#8BC34A"
            btn_danger_bg = "#EF5350"
            btn_primary_bg = "#64B5F6"
            btn_info_bg = "#4DD0E1"

            self.theme_toggle_button.config(text="حالت تاریک")

        self.root.config(bg=bg_color)

        # ✅ اصلاح ۱: رنگ متن دکمه‌ها بر اساس تم
        text_color = "white" if theme_name == "dark" else "black"

        # استایل‌دهی برای ttk widgets
        self.style.configure(".", font=(GLOBAL_FONT_NAME, GLOBAL_FONT_SIZE), background=bg_color, foreground=fg_color)
        self.style.configure("TFrame", background=bg_color)
        self.style.configure("TLabelframe", background=bg_color, foreground=fg_color)
        self.style.configure("TLabelframe.Label", background=bg_color, foreground=fg_color,
                             font=(GLOBAL_FONT_NAME, GLOBAL_FONT_SIZE, "bold"))
        self.style.configure("TLabel", background=bg_color, foreground=fg_color)
        self.style.configure("TEntry", fieldbackground=entry_bg, foreground=entry_fg, borderwidth=1, relief="solid")
        self.style.configure("TCombobox", fieldbackground=entry_bg, foreground=entry_fg, selectbackground=select_bg,
                             selectforeground=select_fg, borderwidth=1, relief="solid")
        self.style.configure("TButton", background=btn_default_bg, foreground=fg_color, borderwidth=1, relief="raised")
        self.style.map("TButton", background=[("active", btn_hover_bg)])

        self.style.configure("Statusbar.TLabel", background=bg_color, foreground=fg_color,
                             font=(GLOBAL_FONT_NAME, GLOBAL_FONT_SIZE))

        # ✅ اصلاح رنگ متن دکمه‌ها در هر تم
        self.style.configure("Accent.TButton", background=btn_accent_bg, foreground=text_color)
        self.style.map("Accent.TButton", background=[("active", self.darken_color(btn_accent_bg, 20))])

        self.style.configure("Danger.TButton", background=btn_danger_bg, foreground=text_color)
        self.style.map("Danger.TButton", background=[("active", self.darken_color(btn_danger_bg, 20))])

        self.style.configure("Primary.TButton", background=btn_primary_bg, foreground=text_color)
        self.style.map("Primary.TButton", background=[("active", self.darken_color(btn_primary_bg, 20))])

        self.style.configure("Info.TButton", background=btn_info_bg, foreground=text_color)
        self.style.map("Info.TButton", background=[("active", self.darken_color(btn_info_bg, 20))])

        # استایل برای Treeview
        self.style.configure("Treeview", background=entry_bg, foreground=fg_color, fieldbackground=entry_bg)
        self.style.configure("Treeview.Heading", background=btn_default_bg, foreground=fg_color,
                             font=(GLOBAL_FONT_NAME, GLOBAL_FONT_SIZE, "bold"))
        self.style.map("Treeview.Heading", background=[("active", btn_hover_bg)])

        if theme_name == "dark":
            self.tree.tag_configure("tag_red", background="#8b0000", foreground="black")
            self.tree.tag_configure("tag_green", background="#006400", foreground="black")
            self.tree.tag_configure("tag_yellow", background="#b8860b", foreground="black")
            self.tree.tag_configure("tag_blue", background="#00008b", foreground="black")
            self.tree.tag_configure("alternate_row", background="#3a3a3a", foreground=fg_color)
        else:  # light
            self.tree.tag_configure("tag_red", background="#f8d7da", foreground="black")
            self.tree.tag_configure("tag_green", background="#d4edda", foreground="black")
            self.tree.tag_configure("tag_yellow", background="#fff3cd", foreground="black")
            self.tree.tag_configure("tag_blue", background="#d1ecf1", foreground="black")
            self.tree.tag_configure("alternate_row", background="#e0e0e0", foreground=fg_color)

        # ✅ اطمینان از رنگ‌بندی درست Text در هر تم
        self.entries["description"].config(bg=entry_bg, fg=entry_fg, insertbackground=entry_fg)

        # 🔁 بازنویسی تگ‌های Treeview
        self.style_treeview_tags()

    def style_treeview_tags(self):
        """اعمال رنگ‌بندی به تگ‌های Treeview بر اساس تم فعلی."""
        for i, item_id in enumerate(self.tree.get_children()):
            tags = list(self.tree.item(item_id, "tags"))
            tags = [t for t in tags if t not in ["alternate_row"]]  # حذف تگ قبلی

            if i % 2 == 0:
                tags.append("alternate_row")

            self.tree.item(item_id, tags=tags)

    def toggle_theme(self):
        """تغییر تم بین حالت روشن و تاریک"""
        if self.current_theme == "light":
            messagebox.showinfo("حالت تاریک غیرفعال است", "فعلاً فقط حالت روشن فعال است. به‌زودی اضافه خواهد شد.")
        else:
            self.apply_theme("light")
        self.config["theme"] = self.current_theme
        save_config(self.config)
        self.update_status_bar(f"تم به حالت {'تاریک' if self.current_theme == 'dark' else 'روشن'} تغییر یافت.")

    def darken_color(self, hex_color, percent):
        """تیره کردن یک رنگ هگز دسیمال"""
        hex_color = hex_color.lstrip('#')
        rgb = tuple(int(hex_color[i:i + 2], 16) for i in (0, 2, 4))

        darkened_rgb = []
        for val in rgb:
            darkened_val = int(val * (100 - percent) / 100)
            darkened_rgb.append(min(255, max(0, darkened_val)))

        return '#%02x%02x%02x' % tuple(darkened_rgb)

    def export_to_excel(self):
        """خروجی به فایل Excel"""
        if not EXCEL_AVAILABLE:
            messagebox.showerror("خطا", "کتابخانه openpyxl نصب نیست.")
            return

        if not self.data:
            messagebox.showinfo("اطلاع", "هیچ داده‌ای برای خروجی وجود ندارد.")
            return

        filepath = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            title="ذخیره فایل اکسل"
        )
        if not filepath:
            return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "گزارش پروژه‌ها"

            headers = ["نام مهندس", "آدرس", "متراژ", "تعداد اتاق", "تاریخ ویزیت",
                       "تاریخ تماس بعدی", "وضعیت", "توضیحات", "تاریخ پایان"]
            ws.append(headers)

            ws.sheet_view.rightToLeft = True

            fill_red = PatternFill(start_color="f8d7da", end_color="f8d7da", fill_type="solid")
            fill_green = PatternFill(start_color="d4edda", end_color="d4edda", fill_type="solid")
            fill_yellow = PatternFill(start_color="fff3cd", end_color="fff3cd", fill_type="solid")
            fill_blue = PatternFill(start_color="d1ecf1", end_color="d1ecf1", fill_type="solid")

            for rec in self.data:
                row_data = [
                    rec.get("name", ""),
                    rec.get("address", ""),
                    rec.get("area", ""),
                    rec.get("rooms", ""),
                    rec.get("visit_date", ""),
                    rec.get("next_call_date", ""),
                    rec.get("status", ""),
                    rec.get("description", ""),
                    rec.get("end_date", "")
                ]
                ws.append(row_data)

                fill = None
                status = rec.get("status", "")
                if status == "از دست رفته":
                    fill = fill_red
                elif status == "خرید":
                    fill = fill_green
                elif status == "انتظار":
                    fill = fill_yellow
                elif status == "در انتظار تماس مجدد":
                    fill = fill_blue

                if fill:
                    for cell in ws[ws.max_row]:
                        cell.fill = fill

            for col_idx, column in enumerate(ws.columns, 1):
                max_length = 0
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

            wb.save(filepath)
            messagebox.showinfo("موفق", f"فایل Excel با موفقیت در \n{filepath}\nذخیره شد.")
            self.update_status_bar("فایل Excel با موفقیت ذخیره شد.")

        except Exception as e:
            messagebox.showerror("خطا", f"خطا در ایجاد فایل Excel: {str(e)}")
            self.update_status_bar("خطا در ذخیره فایل Excel.")

    def export_to_pdf(self):
        """خروجی به فایل PDF"""
        if not PDF_AVAILABLE:
            messagebox.showerror("خطا", "کتابخانه reportlab نصب نیست.")
            return

        if not self.data:
            messagebox.showinfo("اطلاع", "هیچ داده‌ای برای خروجی وجود ندارد.")
            return

        filepath = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")],
            title="ذخیره فایل PDF"
        )
        if not filepath:
            return

        if not os.path.exists(PDF_FONT_PATH):
            messagebox.showerror("خطای فونت PDF",
                                 "فایل فونت فارسی برای PDF یافت نشد. "
                                 "لطفاً فایل Tanha.ttf را دانلود کرده و کنار برنامه قرار دهید."
                                 "\n(لینک دانلود در توضیحات داده شده است)")
            return

        try:
            c = canvas.Canvas(filepath, pagesize=A4)
            width, height = A4
            margin = 2 * cm
            y = height - margin

            c.setFont(PDF_FONT_NAME, 16)
            c.drawRightString(width - margin, y, "گزارش مدیریت پروژه‌ها")
            y -= 1.5 * cm

            c.setFont(PDF_FONT_NAME, 8)
            current_greg_time = datetime.now()
            current_shamsi_time = gregorian_datetime_to_shamsi_str(current_greg_time)
            c.drawRightString(width - margin, y,
                              f"تاریخ تولید: {current_shamsi_time} {current_greg_time.strftime('%H:%M')}")
            y -= 2 * cm

            c.setFont(PDF_FONT_NAME, 9)
            headers = ["تاریخ پایان", "توضیحات", "وضعیت", "تماس بعدی", "تاریخ ویزیت", "اتاق", "متراژ", "آدرس",
                       "نام مهندس"]
            col_widths = [2.5 * cm, 4 * cm, 2.5 * cm, 2.5 * cm, 2.5 * cm, 1 * cm, 1.5 * cm, 4 * cm, 2.5 * cm]

            x_start = width - margin
            current_x = x_start
            for i, header in enumerate(headers):
                text_width = pdfmetrics.stringWidth(header, PDF_FONT_NAME, 9)
                current_x -= col_widths[i]
                c.drawString(current_x + (col_widths[i] - text_width) / 2, y, header)
            y -= 0.5 * cm

            c.line(margin, y, width - margin, y)
            y -= 0.5 * cm

            c.setFont(PDF_FONT_NAME, 8)
            row_height = 0.8 * cm

            for rec in self.data:
                if y < margin + 2 * cm:
                    c.showPage()
                    c.setFont(PDF_FONT_NAME, 9)
                    y = height - margin - 2 * cm
                    current_x = x_start
                    for i, header in enumerate(headers):
                        text_width = pdfmetrics.stringWidth(header, PDF_FONT_NAME, 9)
                        current_x -= col_widths[i]
                        c.drawString(current_x + (col_widths[i] - text_width) / 2, y + 1.5 * cm, header)
                    c.line(margin, y + 1 * cm, width - margin, y + 1 * cm)
                    y -= 0.5 * cm
                    c.setFont(PDF_FONT_NAME, 8)

                status = rec.get("status", "")
                fill_color = colors.white
                if status == "از دست رفته":
                    fill_color = colors.HexColor("#f8d7da")
                elif status == "خرید":
                    fill_color = colors.HexColor("#d4edda")
                elif status == "انتظار":
                    fill_color = colors.HexColor("#fff3cd")
                elif status == "در انتظار تماس مجدد":
                    fill_color = colors.HexColor("#d1ecf1")

                c.setFillColor(fill_color)
                c.rect(margin, y - 0.2 * cm, sum(col_widths), row_height, fill=1, stroke=0)
                c.setFillColor(colors.black)

                values_to_print = [
                    rec.get("end_date", ""),
                    rec.get("description", "")[:25] + "..." if len(rec.get("description", "")) > 25 else rec.get(
                        "description", ""),
                    rec.get("status", ""),
                    rec.get("next_call_date", ""),
                    rec.get("visit_date", ""),
                    rec.get("rooms", ""),
                    rec.get("area", ""),
                    rec.get("address", "")[:25],
                    rec.get("name", "")[:15]
                ]

                current_x = x_start
                for i, value in enumerate(values_to_print):
                    text_width = pdfmetrics.stringWidth(str(value), PDF_FONT_NAME, 8)
                    current_x -= col_widths[i]
                    c.drawString(current_x + (col_widths[i] - text_width) / 2, y, str(value))

                y -= row_height

            c.save()
            messagebox.showinfo("موفق", f"فایل PDF با موفقیت در \n{filepath}\nذخیره شد.")
            self.update_status_bar("فایل PDF با موفقیت ذخیره شد.")

        except Exception as e:
            messagebox.showerror("خطا", f"خطا در ایجاد فایل PDF: {str(e)}")
            if "Cannot find TrueType font file" in str(e):
                messagebox.showerror("خطای فونت PDF",
                                     "فایل فونت فارسی برای PDF یافت نشد. "
                                     "لطفاً فایل Tanha.ttf را دانلود کرده و کنار برنامه قرار دهید.")
            self.update_status_bar("خطا در ذخیره فایل PDF.")


def main():
    """تابع اصلی برنامه"""
    root = tk.Tk()
    app = ProjectManager(root)

    root.protocol("WM_DELETE_WINDOW", lambda: (save_data(app.data), save_config(app.config), root.destroy()))

    try:
        root.mainloop()
    except KeyboardInterrupt:
        save_data(app.data)
        save_config(app.config)
        root.destroy()


if __name__ == "__main__":
    main()
